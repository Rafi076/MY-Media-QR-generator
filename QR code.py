
# # img = qr.make("https://www.facebook.com/rafio076?mibextid=ZbWKwL")
# # img.save("Facebook qr.png")



# Here, the png img will save in QR code genarator folder

# import qrcode as qr
# from openpyxl import load_workbook
# workbook = load_workbook("QR code.xlsx") # load the excel workbook

# #select the active worksheet
# sheet = workbook.active

# # Create and modify cell value
# img = qr.make(sheet['A2'].value)
# sheet['E2'] = img.save("Facebook qr.png")

# img = qr.make(sheet['B2'].value)
# sheet['F2'] = img.save("Linkedin qr.png")

# img = qr.make(sheet['C2'].value)
# sheet['G2'] = img.save("GitHub qr.png")

# #save the workbook
# workbook.save('QR code.xlsx')




# # Here, the png img will save in Excel sheet
import qrcode as qr
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.units import points_to_pixels

# Function to convert inches to points
def inches_to_points(inches):
    return inches * 72  # 1 inch = 72 points

# Load the Excel workbook
workbook = load_workbook("QR code.xlsx")

# Select the active worksheet
sheet = workbook.active

# Function to create and insert QR code image into Excel cell
def create_and_insert_qr_code(url, cell):
    # Making QR
    img = qr.make(url)
    
    # Create a unique name for each image
    img_path = f"{url.split('/')[-1]}_qr.png"  
    img.save(img_path)
    
    # Create Image object and set dimensions
    img = Image(img_path)
    img.height = inches_to_points(0.99)
    img.width = inches_to_points(0.99)
    
    # Insert image into Excel cell
    sheet.add_image(img, cell)  

# Create and insert QR codes for Facebook, LinkedIn, and GitHub URLs
create_and_insert_qr_code(sheet['A2'].value, "E2")
create_and_insert_qr_code(sheet['B2'].value, "F2")
create_and_insert_qr_code(sheet['C2'].value, "G2")

# Save the workbook
workbook.save('QR code.xlsx')
